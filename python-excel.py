import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

wb = load_workbook('公有停車場.xlsx')
sheet = wb['工作表1']
cell = sheet.cell(row=1, column=11)
print(sheet.max_row)

def get_Longitude_Latitude(addr):
    respoense = requests.get("https://www.google.com/maps/place?q=" + addr)
    soup = BeautifulSoup(respoense.text, "html.parser")
    text = soup.prettify()
    print(text)
    initial_pois = text.find("https://maps.google.com/maps/api/staticmap?center=")
    data = text[initial_pois+50:initial_pois+76]
    Latitude, Longitude  = data.split("%2C")
    return Latitude, Longitude

for i in range(2,sheet.max_row):
    print(i)
    addr = sheet.cell(row=i, column=4).value
    Latitude, Longitude = get_Longitude_Latitude(addr)
    print("addr = ", addr)
    print("Latitude = ", Latitude)
    print("Longitude = ", Longitude)
    excel_Latitude = sheet.cell(row=i, column=11)
    excel_Longitude = sheet.cell(row=i, column=12)
    excel_Latitude.value = Latitude
    excel_Longitude.value = Longitude
wb.save("test1.xlsx")
